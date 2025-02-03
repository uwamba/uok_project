
from pyexpat.errors import messages
from .serializers import TestSerializer
from django.shortcuts import render, redirect
from .forms import TestForm
from .forms import QuestionForm, QuestionOptionFormSet
from django.shortcuts import render, redirect, get_object_or_404
from .models import Subject, Test, Question, QuestionOption
from results.models import Result, ResultDetail
from users.models import Admin
from .forms import TestForm, QuestionFormSet, QuestionOptionFormSet
from datetime import timedelta, datetime
from .forms import TestForm
from dal import autocomplete
from .models import Candidate
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from .forms import CandidateForm, UserRegistrationForm

from django.contrib.auth import get_user_model
from .models import Candidate, Company

from django.http import HttpResponse
from .forms import ImportQuestionsForm
from .models import Question, QuestionOption
from import_export import resources
from .resources import QuestionResource, QuestionOptionResource
import openpyxl
from io import BytesIO
from openpyxl import load_workbook
import random
from faker import Faker
from datetime import timedelta


def test_detail(request, test_id,result_id):
    test = get_object_or_404(Test, id=test_id)
    results= Result.objects.get(id=result_id)
    #results_answer=results.details.all()
    # For a list of values of a specific field (e.g., 'selected_option_id')
    results_answers = ResultDetail.objects.filter(result_id=result_id)
    results_answer = ResultDetail.objects.filter(result_id=result_id).values('selected_option_id')
    selected_option_ids = [item['selected_option_id'] for item in results_answer if item['selected_option_id'] is not None]

    print(results_answers.values('question_id','answer_text'))
    questions = test.questions.all()
    
    context = {
        'test': test,
        'questions': questions,
        'results_answer':results_answer,
        'results_answers':results_answers,
        'selected_option_ids': selected_option_ids,
    }
    return render(request, 'test_details.html', context)

def create_test(request):
    if request.method == 'POST':
        form = TestForm(request.POST, user=request.user)
        if form.is_valid():
            form.save()  # Save the new test to the database
            return redirect('admin_dashboard')  # Redirect to a list or confirmation page after saving
    else:
        print('request user', request.user.id)  
        form = TestForm(user=request.user.id)  # Pass the logged-in user to the form

    return render(request, 'create_test.html', {'form': form})


def add_test(request):
    if request.method == 'POST':
        # Save Test
        duration_minutes = int(request.POST.get('duration'))
       # admin = Admin.objects.get(user=request.user)  # Get the Admin object linked to the logged-in user
    
        # Get the company related to the logged-in admin
        #company = admin.company
        test = Test.objects.create(
            title=request.POST.get('test_title'),
            subject_id=request.POST.get('subject'),
            start_time = datetime.strptime(request.POST.get('start_time'), '%Y-%m-%dT%H:%M'),
            end_time = datetime.strptime(request.POST.get('end_time'), '%Y-%m-%dT%H:%M'),

            total_marks=request.POST.get('total_marks'),
            max_attempts=request.POST.get('max_attempts'),
            duration = timedelta(minutes=duration_minutes),
            created_by=request.user
        )

        # Save Questions and Options
        questions = request.POST.getlist('question_text')
        for i, question_text in enumerate(questions):
            # Get the image from the request (it will be stored in request.FILES)
            image = request.FILES.get(f'image_{i}', None)  # Adjust the field name as per the form field name for image input

            # Create the question and save to the database
            question = Question.objects.create(
                test=test,
                text=question_text,
                marks=request.POST.getlist('question_marks')[i],
                question_type=request.POST.getlist('question_type')[i],
                max_selection=request.POST.getlist('max_selection')[i],
                answer_text=request.POST.get(f'answer_text_{i}', None),  # Handle text answers
                image=image  # Save the image if it's provided
            )


            # Save Options if the question is not a text question
            if question.question_type != 'text':
                options = request.POST.getlist(f'option_text_{i}')
                correct_options = request.POST.getlist(f'is_correct_{i}')
                for j, option_text in enumerate(options):
                    QuestionOption.objects.create(
                        question=question,
                        text=option_text,
                        is_correct=(str(j) in correct_options)
         
                    )

        selected_candidates = request.POST.getlist('candidates')  # Get the list of selected candidate IDs
        for candidate_id in selected_candidates:
            candidate = Candidate.objects.get(id=candidate_id)
            test.candidates.add(candidate)  # Associate the candidate with the test
        return redirect('admin_dashboard')
    form = TestForm()
    candidates = Candidate.objects.all() 
    return render(request, 'add_test.html', {'subjects': Subject.objects.filter(company=request.user.admin.company),'form': form,'candidates': candidates})
from datetime import datetime, timedelta
from django.shortcuts import render, get_object_or_404, redirect
from .models import Test, Question, QuestionOption, Candidate, Subject
from .forms import TestForm

def edit_test(request, test_id):
    test = get_object_or_404(Test, id=test_id)

    if request.method == 'POST':
        # Update test details
        duration_minutes = int(request.POST.get('duration'))
        test.title = request.POST.get('test_title')
        test.subject_id = request.POST.get('subject')
        test.start_time = datetime.strptime(request.POST.get('start_time'), '%Y-%m-%dT%H:%M')
        test.end_time = datetime.strptime(request.POST.get('end_time'), '%Y-%m-%dT%H:%M')
        test.total_marks = request.POST.get('total_marks')
        test.max_attempts = request.POST.get('max_attempts')
        test.duration = timedelta(minutes=duration_minutes)
        test.save()

        # Update Questions
        # Update Questions
        questions = request.POST.getlist('question_text')

        for i, question_text in enumerate(questions):
            image = request.FILES.get(f'image_{i}', None)  # Dynamically get the image for each question

            # Check if the question already exists (in case of an update)
            question_id = None
            if 'question_id' in request.POST and len(request.POST.getlist('question_id')) > i:
                question_id = request.POST.getlist('question_id')[i]  # Only get the question_id if it exists

            if question_id:  # If updating an existing question
                print('question_id', question_id)
                question = Question.objects.get(id=question_id)
                
                if not image:  # No new image uploaded
                    image = question.image  # Retain the existing image

                # Update question fields
                question.text = question_text
                question.marks = request.POST.getlist('question_marks')[i]
                question.question_type = request.POST.getlist('question_type')[i]
                question.max_selection = request.POST.getlist('max_selection')[i]
                question.answer_text = request.POST.get(f'answer_text_{i}', None)
                question.image = image  # Keep old image if no new one is uploaded

            else:  # Creating a new question
                question = Question(
                    test=test,
                    text=question_text,
                    marks=request.POST.getlist('question_marks')[i],
                    question_type=request.POST.getlist('question_type')[i],
                    max_selection=request.POST.getlist('max_selection')[i],
                    answer_text=request.POST.get(f'answer_text_{i}', None),
                    image=image if image else None  # Set image if provided
                )

            question.save()  # Save the question after modifications

            # Save options if the question is not a text type
            if question.question_type != 'text':
                question.options.all().delete()
                options = request.POST.getlist(f'option_text_{i}')
                correct_options = request.POST.getlist(f'is_correct_{i}')
                for j, option_text in enumerate(options):
                    QuestionOption.objects.create(
                        question=question,
                        text=option_text,
                        is_correct=(str(j) in correct_options)
                    )


        # Update assigned candidates
        selected_candidates = request.POST.getlist('candidates')
        test.candidates.clear()  # Remove old candidates
        for candidate_id in selected_candidates:
            candidate = Candidate.objects.get(id=candidate_id)
            test.candidates.add(candidate)

        return redirect('admin_dashboard')

    # Pre-fill the form with existing data
    form = TestForm(instance=test)
    candidates = Candidate.objects.filter(company=test.created_by.admin.company)
    questions = Question.objects.filter(test=test)
    subjects = Subject.objects.filter(company=test.created_by.admin.company)

    if test.duration:
       duration_minutes = int(test.duration.total_seconds() // 60)
    else:
        duration_minutes = 0

    return render(request, 'edit_test.html', {
        'test': test,
        'form': form,
        'candidates': candidates,
        'questions': questions,
        'duration_minutes': duration_minutes, 
        'subjects': subjects,  
    })



class CandidateAutocomplete(autocomplete.Select2QuerySetView):
    def get_queryset(self):
        # Optionally filter candidates based on the search query
        qs = Candidate.objects.all()
        if self.q:
            qs = qs.filter(full_name__icontains=self.q)  # Search by full name
        return qs
def create_question(request):
    if request.method == 'POST':
        question_form = QuestionForm(request.POST, request.FILES)
        option_formset = QuestionOptionFormSet(request.POST)
        if question_form.is_valid() and option_formset.is_valid():
            question = question_form.save()
            option_formset.instance = question
            option_formset.save()
            return redirect('success_page')  # Replace with your success URL
    else:
        question_form = QuestionForm()
        option_formset = QuestionOptionFormSet()

    return render(request, 'create_question.html', {
        'question_form': question_form,
        'option_formset': option_formset
    })
    
from .forms import SubjectForm

def create_subject(request):
    # Get the company associated with the logged-in admin
    company = request.user.admin.company  # Assuming the admin has a related 'Company' via the 'Admin' model

    if request.method == 'POST':
        form = SubjectForm(request.POST)
        if form.is_valid():
            # Set the company automatically to the logged-in admin's company
            subject = form.save(commit=False)
            subject.company = company  # Assign the company
            subject.save()

            #messages.success(request, "Subject created successfully!")
            return redirect('admin_dashboard')  # Redirect to the subject list page
    else:
        form = SubjectForm()

    return render(request, 'create_subject.html', {'form': form})



@login_required
def register_candidate(request):
    if request.method == "POST":
        username = request.POST.get('username')
        password = request.POST.get('password')
        full_name = request.POST.get('full_name')
        date_of_birth = request.POST.get('date_of_birth')
        phone_number = request.POST.get('phone_number')

        # Create a new user using the custom user model
        User = get_user_model()
        user = User.objects.create_user(username=username, password=password)

        # Set user's full name and other information
        user.first_name = full_name
        user.save()

        # Get the company of the logged-in admin
        company = request.user.admin.company 
        
        # Create a new candidate
        Candidate.objects.create(
            user=user,
            company=company,
            date_of_birth=date_of_birth,
            phone_number=phone_number
        )

        return redirect('admin_dashboard')  # Redirect to a success page or any other page

    return render(request, 'register_candidate.html')
@login_required
def list_candidates(request):
    # Get the company of the logged-in admin
    company = request.user.admin.company  # Assuming logged-in admin has a Company linked to them
    
    # Fetch all candidates associated with that company
    candidates = Candidate.objects.filter(company=company)
    
    # Pass the list of candidates to the template
    return render(request, 'list_candidates.html', {'candidates': candidates})
@login_required
def candidate_detail(request, candidate_id):
    candidate = get_object_or_404(Candidate, id=candidate_id)
    return render(request, 'candidate_detail.html', {'candidate': candidate})
@login_required
def candidate_edit(request, candidate_id):
    candidate = get_object_or_404(Candidate, id=candidate_id)

    if request.method == 'POST':
        form = CandidateForm(request.POST, instance=candidate)
        if form.is_valid():
            form.save()
            return redirect('candidate_detail', candidate_id=candidate.id)  # Redirect to the candidate detail page
    else:
        form = CandidateForm(instance=candidate)

    return render(request, 'candidate_edit.html', {'form': form, 'candidate': candidate})

faker = Faker()

def random_test_data(request):
    # Generate random test data using Faker
    title = faker.sentence(nb_words=3)  # Random title (e.g., "Math Test")
    subject = random.choice(Subject.objects.all())  # Randomly choose a subject from existing subjects
    start_time = faker.date_this_year()  # Random start time in this year
    end_time = start_time + timedelta(hours=random.randint(1, 3))  # Random end time (between 1 and 3 hours later)
    total_marks = random.randint(50, 100)  # Random total marks between 50 and 100
    max_attempts = random.randint(1, 3)  # Random number of attempts between 1 and 3
    candidates = random.sample(list(Candidate.objects.all()), k=random.randint(1, 5))  # Random candidates
    duration_type = random.choice([Test.TEST_DURATION, Test.QUESTION_DURATION])  # Randomly choose duration type
    duration = timedelta(minutes=random.randint(30, 120))  # Random duration between 30 to 120 minutes
    
    # Create the Test object with random values
    test = Test.objects.create(
        title=title,
        subject=subject,
        start_time=start_time,
        end_time=end_time,
        total_marks=total_marks,
        max_attempts=max_attempts,
        duration=duration,
        counterType=duration_type,
         created_by=request.user
    )
    
    # Assign random candidates
    test.candidates.set(candidates)
    test.save()

    return test

def import_questions(request):
    if request.method == 'POST':
        form = ImportQuestionsForm(request.POST, request.FILES)
        if form.is_valid():
            # Open the uploaded Excel file
            excel_file = request.FILES['file']
            wb = openpyxl.load_workbook(excel_file)
            sheet = wb.active

            # Randomize test data
            test = random_test_data(request)  # Randomize and create test

            for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip header row
                question_text = row[0]  # Question text
                marks = row[1]          # Marks
                question_type = row[2]  # Question type (single, multiple, text)

                # Create the Question object and link it to the random Test
                question = Question.objects.create(
                    test=test,
                    text=question_text,
                    marks=marks,
                    question_type=question_type,
                )

                # Process options and create QuestionOption objects
                for i in range(3, len(row), 2):  # Start from Option 1, skip every other column
                    option_text = row[i]
                    is_correct = row[i+1] == "True"  # Convert "True" to boolean

                    # Check if option_text is not empty before creating the QuestionOption
                    if option_text:
                        # Create the QuestionOption object only if option_text is not empty
                        QuestionOption.objects.create(
                            question=question,
                            text=option_text,
                            is_correct=is_correct
                        )
                    else:
                        # Option text is empty, you could log or handle it differently
                        print(f"Skipping empty option at row {row[0]}")  # Logging empty option
                        continue

            return redirect('admin_dashboard')  # Redirect to a page that shows the questions
    else:
        form = ImportQuestionsForm()

    return render(request, 'import_questions.html', {'form': form})
def export_questions(request):
    # Create an in-memory workbook
    wb = openpyxl.Workbook()
    sheet = wb.active

    # Create headers
    sheet.append([
        'Question Text', 'Marks', 'Question Type', 
        'Option 1', 'Is Correct 1', 'Option 2', 'Is Correct 2', 
        'Option 3', 'Is Correct 3', 'Option 4', 'Is Correct 4'
    ])

    # Add questions and options
    questions = Question.objects.all()
    for question in questions:
        row = [
            question.text, question.marks, question.question_type
        ]
        
        # Add options for each question
        options = question.options.all()
        for option in options:
            row.append(option.text)
            row.append('True' if option.is_correct else 'False')
        
        # Ensure there are 4 options (if there are less, add blank ones)
        while len(row) < 10:  # 4 options * 2 columns (text and is_correct)
            row.append('')
        
        sheet.append(row)

    # Save the file to an HTTP response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="questions.xlsx"'

    # Save the workbook to the response
    wb.save(response)
    return response
